import time,os
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
import re
import uiautomation as auto
from selenium.webdriver.chrome.options import Options

class auto_case:
    def __init__(self):
        # self.options = Options()
        # self.options.add_argument('--headless')
        self.driver = webdriver.Chrome()
        #self.driver = webdriver.Chrome(chorme_path,chorme_options=options)
        self.excel = openpyxl.load_workbook(r'C:\testplan.xlsx')
        self.imformation_sheet = self.excel['information']
        self.user = self.imformation_sheet['B1'].value
        self.password = self.imformation_sheet['B2'].value
        self.project = self.imformation_sheet['B3'].value
        self.testplan = self.imformation_sheet['B4'].value
        self.project_issue_id=self.imformation_sheet['B5'].value
        self.case_sheet = self.excel['case']
        self.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # 加线框
        self.font = Font(size=12, name='Calibri')  # 设置样式
        self.align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 设置对齐方式

    def enter_plan(self):
        self.driver.get("http://tdms.lenovo.com/tdms/loginAction!login.do")
        self.driver.maximize_window()
        element = WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.ID, 'username')))#最长等待1000s
        element.send_keys(self.user)
        self.driver.find_element(By.XPATH, "//input[@id='password']").send_keys(self.password)
        self.driver.find_element(By.XPATH, "//input[@id='loginsub']").click()
        element_test = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'menu_system_test')))
        element_test.click()
        element_excutetest = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'test_menu_executetestplan')))
        element_excutetest.click()
        WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'tableId')))#先确保prjoject name 栏位有出现
        self.driver.find_element(By.LINK_TEXT, self.project).click()
        WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'tableId')))#先确保prjoject name 栏位有出现
        time.sleep(2)
        self.driver.find_element(By.LINK_TEXT, self.testplan).click()
        WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'caseId')))  # 到这里为止是进入了test plan , 下面就是plan下所有的test case
        time.sleep(2)
    def enter_testcase(self):
        main_window=self.driver.current_window_handle#定位到当前窗口
        all_rows = self.case_sheet.max_row#获取最大行数
        for y in range(2, all_rows + 1):#如果是4行，要除去表头title栏位，所以要从第二行开始，（2.5）的取值刚好是234行，所以要加1
            every_row = []
            for cell in self.case_sheet[y]:#case sheet的第一行
              every_row.append(cell.value)
            try:
                self.driver.find_element(By.LINK_TEXT, every_row[0]).click()#找到case ID点击,如果这里报错就是没找到case ID
            except:
                target_cell = self.case_sheet.cell(column=4, row=y)
                target_cell.border = self.border
                target_cell.alignment = self.align
                target_cell.font = self.font
                target_cell.value = "this case doesn't exit in TDMS plan"
                self.excel.save(r'C:\testplan.xlsx')
                continue#用不用continue关键在于下面的代码有没有必要继续运行，还是这里报错了, 后面就没必要运行了，继续开始的循环就必须加continue, 不加会接着后面的代码继续运行
            popup_windows = self.driver.window_handles[1]  # 窗口是列表[0,1,2]1就表示第二哥窗口，-1表示最后一个
            self.driver.switch_to.window(popup_windows)#定位到新窗口
            try:
                WebDriverWait(self.driver, 1000).until(ec.presence_of_element_located((By.XPATH, "//*[@id='excuteCaseTableLeft']/thead/tr/td[3]/strong/font")))#等待新加载的页面出现test case item元素代表加载完了再进行下一步
                if every_row[1] == None and every_row[2]==None:#Pass且无需要log的情况
                    workloading_time=self.driver.find_elements(By.XPATH,"//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[1]/td[1]/span")#找到所有section包含workloding时间的标签
                    workload_list=[]
                    for each_workloading_time in workloading_time:
                        text=each_workloading_time.get_attribute('textContent')#获取标签的内容例如Microsoft Windows ML (240)，240就是时间
                        time_list=re.findall('\((.*?)\)', text)#匹配()里面的240，\(.*?\)这种会把（）也匹配出来了，返回(240)， （）在正则表达式本身就有含义，表示要取的内容。所以加转义符取括号里面的内容，返回列表[240]
                        for x in range(0,len(time_list)):
                            try:
                                int_change=int(time_list[x])
                                workload_list.append(int_change)#到这里为止workload list里就装满了时间[120,240]这样子
                            except:
                                pass
                    input_time=self.driver.find_elements(By.XPATH,"//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[2]/td[1]/input[2]")#找到所有需要输入workloading的标签
                    #input标签长度和workload list 肯定一样匹配的，所以下面的方法巧妙解决一一对应填上
                    i=0
                    for eachinput in input_time:
                        content = eachinput.get_attribute('value')#获取之前有没有填过值
                        if content!='':#有值就不填且加1到下一个
                            i=i+1
                        else: #等于空就要填值
                            eachinput.send_keys(workload_list[i])
                            i=i+1#到这里截止每个workloading就填好了
                    steps= self.driver.find_elements(By.XPATH, "//table[@id='excuteCaseTableRight']/tbody/tr/td/select")#找到所有步骤
                    for everystep in steps:
                        step = Select(everystep)
                        step.select_by_value('2')#2代表pass
                        time.sleep(0.1)#打快了虽然每一步都能打上，但最后结果部显示100%, 经过实验加0.1就可以了
                    self.driver.find_element(By.XPATH, "//input[@name='Save']").click()
                    time.sleep(2)
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    target_cell = self.case_sheet.cell(column=4, row=y)
                    target_cell.border = self.border
                    target_cell.alignment = self.align
                    target_cell.font = self.font
                    target_cell.value = "Done"
                    self.excel.save(r'C:\testplan.xlsx')
                elif every_row[1] == None and every_row[2] != None:  # Pass且需要log的情况
                    workloading_time = self.driver.find_elements(By.XPATH, "//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[1]/td[1]/span")  # 找到所有section包含workloding时间的标签
                    workload_list = []
                    for each_workloading_time in workloading_time:
                        text = each_workloading_time.get_attribute('textContent')  # 获取标签的内容例如Microsoft Windows ML (240)，240就是时间
                        time_list = re.findall('\((.*?)\)', text)  # 匹配()里面的240，\(.*?\)这种会把（）也匹配出来了，返回(240)， （）在正则表达式本身就有含义，表示要取的内容。所以加转义符取括号里面的内容，返回列表[240]
                        for x in range(0, len(time_list)):#有时候有case是(KBR)(220)这种格式，就回匹配打字符串，所以转为整数再加入列表。字符串转不了就淘汰了
                            try:
                                int_change = int(time_list[x])
                                workload_list.append(int_change)  # 到这里为止workload list里就装满了时间[120,240]这样子
                            except:
                                pass # 到这里为止workload list里就装满了时间[120,240]这样子
                    input_time = self.driver.find_elements(By.XPATH, "//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[2]/td[1]/input[2]")  # 找到所有需要输入workloading的标签
                    # input标签长度和workload list 肯定一样匹配的，所以下面的方法巧妙解决一一对应填上
                    i = 0
                    for eachinput in input_time:
                        content = eachinput.get_attribute('value')  # 获取之前有没有填过值
                        if content != '':  # 有值就不填且加1到下一个
                            i = i + 1
                        else:  # 等于空就要填值
                            eachinput.send_keys(workload_list[i])
                            i = i + 1  # 到这里截止每个workloading就填好了
                    steps = self.driver.find_elements(By.XPATH, "//table[@id='excuteCaseTableRight']/tbody/tr/td/select")  # 找到所有步骤
                    for everystep in steps:
                        step = Select(everystep)
                        step.select_by_value('2')  # 2代表pass
                        time.sleep(0.1)#打快了虽然每一步都能打上，但最后结果部显示100%, 经过实验加0.1就可以了
                    self.driver.find_element(By.XPATH, "//*[@id='tableTestLogTable']/tbody/tr[1]/td/input").click()  # 找到addbuton点击
                    time.sleep(2)
                    all_window_handles = self.driver.window_handles#['53A92A734DCFB61642699596869E0E7E', '536DFB8531DE61FAA4EBB7860A9A32B3', '6F63829CEB835DA4C3C36CCBD9F9170F']返回3个窗口
                    self.driver.switch_to.window(all_window_handles[2])
                    add = WebDriverWait(self.driver, 60).until(ec.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/input[2]")))  # 等待新窗口的add button出现
                    add.click()
                    time.sleep(3)
                    auto.EditControl(Name="File name:").Click()
                    time.sleep(1)
                    auto.SendKeys("C:\log.zip")
                    time.sleep(1)
                    auto.ButtonControl(ClassName="Button",Name="Open").Click()
                    time.sleep(2)
                    self.driver.find_element(By.XPATH, "/html/body/div/div[3]/input").click()
                    time.sleep(2)
                    all_window_handles = self.driver.window_handles
                    self.driver.switch_to.window(all_window_handles[1])#从新回来第二个窗口
                    self.driver.find_element(By.XPATH, "//input[@name='Save']").click()
                    time.sleep(2)
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    target_cell = self.case_sheet.cell(column=4, row=y)
                    target_cell.border = self.border
                    target_cell.alignment = self.align
                    target_cell.font = self.font
                    target_cell.value = "Done"
                    self.excel.save(r'C:\testplan.xlsx')
                elif every_row[1] != None and every_row[2] != None:  # Fail且需要log的情况
                    workloading_time = self.driver.find_elements(By.XPATH, "//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[1]/td[1]/span")  # 找到所有section包含workloding时间的标签
                    workload_list = []
                    for each_workloading_time in workloading_time:
                        text = each_workloading_time.get_attribute('textContent')  # 获取标签的内容例如Microsoft Windows ML (240)，240就是时间
                        time_list = re.findall('\((.*?)\)', text)  # 匹配()里面的240，\(.*?\)这种会把（）也匹配出来了，返回(240)， （）在正则表达式本身就有含义，表示要取的内容。所以加转义符取括号里面的内容，返回列表[240]
                        for x in range(0, len(time_list)):  # 有时候有case是(KBR)(220)这种格式，就回匹配打字符串，所以转为整数再加入列表。字符串转不了就淘汰了
                            try:
                                int_change = int(time_list[x])
                                workload_list.append(int_change)  # 到这里为止workload list里就装满了时间[120,240]这样子
                            except:
                                pass  # 到这里为止workload list里就装满了时间[120,240]这样子  # 到这里为止workload list里就装满了时间[120,240]这样子
                    input_time = self.driver.find_elements(By.XPATH, "//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[2]/td[1]/input[2]")  # 找到所有需要输入workloading的标签
                    # input标签长度和workload list 肯定一样匹配的，所以下面的方法巧妙解决一一对应填上
                    i = 0
                    for eachinput in input_time:
                        content = eachinput.get_attribute('value')  # 获取之前有没有填过值
                        if content != '':  # 有值就不填且加1到下一个
                            i = i + 1
                        else:  # 等于空就要填值
                            eachinput.send_keys(workload_list[i])
                            i = i + 1  # 到这里截止每个workloading就填好了
                    steps = self.driver.find_elements(By.XPATH, "//table[@id='excuteCaseTableRight']/tbody/tr/td/select")  # 找到所有步骤
                    last_step = Select(steps[-1])#最后一步fail
                    last_step.select_by_value('3')
                    element_comment = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'planTestCaseItemResultComment')))#等待弹框出现
                    self.driver.find_element(By.ID, "testCaseRelatedDefectId1").send_keys(self.project_issue_id)#加link defect ID
                    self.driver.find_element(By.XPATH, "//tbody/tr/td/input[@value='Add' and @onclick='addTestCaseRelatedDefect()']").click()#点add
                    time.sleep(2)
                    self.driver.find_element(By.ID, "planTestCaseItemResultComment").clear()
                    self.driver.find_element(By.ID, "planTestCaseItemResultComment").send_keys("TDMS#" + str(self.project_issue_id) + ":" +every_row[2])
                    time.sleep(1)
                    self.driver.find_element(By.XPATH, "//input[@name='OK']").click()#添加完后点save
                    del steps[-1]#删除刚刚打fail的最后一步
                    for everystep in steps:
                        step = Select(everystep)
                        step.select_by_value('2')  # 2代表pass
                        time.sleep(0.1)#打快了虽然每一步都能打上，但最后结果部显示100%, 经过实验加0.1就可以了
                    self.driver.find_element(By.XPATH, "//*[@id='tableTestLogTable']/tbody/tr[1]/td/input").click()  # 找到addbuton点击
                    time.sleep(2)
                    all_window_handles = self.driver.window_handles#['53A92A734DCFB61642699596869E0E7E', '536DFB8531DE61FAA4EBB7860A9A32B3', '6F63829CEB835DA4C3C36CCBD9F9170F']返回3个窗口
                    self.driver.switch_to.window(all_window_handles[2])
                    add = WebDriverWait(self.driver, 60).until(ec.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/input[2]")))  # 等待新窗口的add button出现
                    add.click()
                    time.sleep(3)
                    auto.EditControl(Name="File name:").Click()
                    time.sleep(1)
                    auto.SendKeys("C:\log.zip")
                    time.sleep(1)
                    auto.ButtonControl(ClassName="Button",Name="Open").Click()
                    time.sleep(2)
                    self.driver.find_element(By.XPATH, "/html/body/div/div[3]/input").click()
                    time.sleep(2)
                    all_window_handles = self.driver.window_handles
                    self.driver.switch_to.window(all_window_handles[1])#从新回来第二个窗口
                    self.driver.find_element(By.XPATH, "//input[@name='Save']").click()
                    time.sleep(2)
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    target_cell = self.case_sheet.cell(column=4, row=y)
                    target_cell.border = self.border
                    target_cell.alignment = self.align
                    target_cell.font = self.font
                    target_cell.value = "Done"
                    self.excel.save(r'C:\testplan.xlsx')
                elif every_row[1] != None and every_row[2] == None:  # Fail不需要log的情况
                    workloading_time = self.driver.find_elements(By.XPATH, "//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[1]/td[1]/span")  # 找到所有section包含workloding时间的标签
                    workload_list = []
                    for each_workloading_time in workloading_time:
                        text = each_workloading_time.get_attribute('textContent')  # 获取标签的内容例如Microsoft Windows ML (240)，240就是时间
                        time_list = re.findall('\((.*?)\)', text)  # 匹配()里面的240，\(.*?\)这种会把（）也匹配出来了，返回(240)， （）在正则表达式本身就有含义，表示要取的内容。所以加转义符取括号里面的内容，返回列表[240]
                        for x in range(0, len(time_list)):  # 有时候有case是(KBR)(220)这种格式，就回匹配打字符串，所以转为整数再加入列表。字符串转不了就淘汰了
                            try:
                                int_change = int(time_list[x])
                                workload_list.append(int_change)  # 到这里为止workload list里就装满了时间[120,240]这样子
                            except:
                                pass  # 到这里为止workload list里就装满了时间[120,240]这样子  # 到这里为止workload list里就装满了时间[120,240]这样子
                    input_time = self.driver.find_elements(By.XPATH, "//*[@id='excuteCaseTableLeft']/tbody/tr[*]/td[3]/table/tbody/tr[2]/td[1]/input[2]")  # 找到所有需要输入workloading的标签
                    # input标签长度和workload list 肯定一样匹配的，所以下面的方法巧妙解决一一对应填上
                    i = 0
                    for eachinput in input_time:
                        content = eachinput.get_attribute('value')  # 获取之前有没有填过值
                        if content != '':  # 有值就不填且加1到下一个
                            i = i + 1
                        else:  # 等于空就要填值
                            eachinput.send_keys(workload_list[i])
                            i = i + 1  # 到这里截止每个workloading就填好了
                    steps = self.driver.find_elements(By.XPATH, "//table[@id='excuteCaseTableRight']/tbody/tr/td/select")  # 找到所有步骤
                    last_step = Select(steps[-1])#最后一步fail
                    last_step.select_by_value('3')
                    element_comment = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'planTestCaseItemResultComment')))#等待弹框出现
                    self.driver.find_element(By.ID, "testCaseRelatedDefectId1").send_keys(self.project_issue_id)#加link defect ID
                    self.driver.find_element(By.XPATH, "//tbody/tr/td/input[@value='Add' and @onclick='addTestCaseRelatedDefect()']").click()#点add
                    time.sleep(2)
                    self.driver.find_element(By.ID, "planTestCaseItemResultComment").clear()
                    self.driver.find_element(By.ID, "planTestCaseItemResultComment").send_keys("TDMS#" + str(self.project_issue_id) + ":" +every_row[2])
                    time.sleep(1)
                    self.driver.find_element(By.XPATH, "//input[@name='OK']").click()#添加完后点save
                    del steps[-1]#删除刚刚打fail的最后一步
                    for everystep in steps:
                        step = Select(everystep)
                        step.select_by_value('2')  # 2代表pass
                        time.sleep(0.1) #打快了虽然每一步都能打上，但最后结果部显示100%, 经过实验加0.1就可以了
                    self.driver.find_element(By.XPATH, "//input[@name='Save']").click()
                    time.sleep(2)
                    self.driver.close()
                    self.driver.switch_to.window(main_window)
                    target_cell = self.case_sheet.cell(column=4, row=y)
                    target_cell.border = self.border
                    target_cell.alignment = self.align
                    target_cell.font = self.font
                    target_cell.value = "Done"
                    self.excel.save(r'C:\testplan.xlsx')
            except:
                target_cell = self.case_sheet.cell(column=4, row=y)
                target_cell.border = self.border
                target_cell.alignment = self.align
                target_cell.font = self.font
                target_cell.value = "new test case format or network problem"
                self.excel.save(r'C:\testplan.xlsx')
                self.driver.close()
                self.driver.switch_to.window(main_window)




    def run(self):
        self.enter_testcase()

if __name__ == "__main__":
        case_go = auto_case()
        case_go.enter_plan()
        case_go.run()







