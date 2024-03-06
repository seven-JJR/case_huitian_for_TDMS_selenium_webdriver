import time,os
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
import openpyxl


class auto_case:
    def __init__(self):
        self.driver = webdriver.Chrome()
        #self.driver = webdriver.Chrome(chorme_path,chorme_options=options)
        self.excel = openpyxl.load_workbook(r'C:\testplan.xlsx')
        self.imformation_sheet = self.excel['information']
        self.user = self.imformation_sheet['B1'].value
        self.password = self.imformation_sheet['B2'].value
        self.project = self.imformation_sheet['B3'].value
        self.testplan = self.imformation_sheet['B4'].value
        self.project_issue_id=self.imformation_sheet['B5'].value
        self.case_sheet = self.excel['case']

    def enter_plan(self):
        self.driver.get("http://tdms.lenovo.com/tdms/loginAction!login.do")
        self.driver.maximize_window()
        element = WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.ID, 'username')))
        element.send_keys(self.user)
        self.driver.find_element(By.XPATH, "//input[@id='password']").send_keys(self.password)
        self.driver.find_element(By.XPATH, "//input[@id='loginsub']").click()
        element_test = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'menu_system_test')))
        element_test.click()
        element_excutetest = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'test_menu_executetestplan')))
        element_excutetest.click()
        element_project = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'tableId')))
        self.driver.find_element(By.LINK_TEXT, self.project).click()
        element_project = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'tableId')))
        self.driver.find_element(By.LINK_TEXT, self.testplan).click()
        element_caseid = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'caseId')))  # 到这里为止是进入了test plan , 下面就是plan下所有的test case
    def enter_testcase(self):
        main_window=self.driver.current_window_handle
        os_not_suuport = ['Win81-64bit', 'Win7- 64bit', 'Win7- 32bit', 'Win10-32bit', 'Win81-32bit']
        all_rows = self.case_sheet.max_row
        for y in range(1, all_rows + 1):
            try:
                every_row = []
                for cell in self.case_sheet[y]:
                    if cell.value is not None:
                        every_row.append(cell.value)
                long = len(every_row)
                if long==1:
                    os_list = []
                    self.driver.find_element(By.LINK_TEXT, every_row[0]).click()
                    popup_windows = self.driver.window_handles[-1]
                    self.driver.switch_to.window(popup_windows)
                    WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH, "//table[@id='excuteCaseTableRight']/thead/tr/td/strong/font")))
                    os = self.driver.find_elements(By.XPATH, "//table[@id='excuteCaseTableRight']/thead/tr/td/strong/font")
                    workloading =self.driver.find_element(By.XPATH, "//input[@id='caseForm_caseBO_workloading']")
                    workloading_text = workloading.get_attribute('value')
                    for everyos in os:
                        text = everyos.text
                        os_list.append(text)
                    list_compare=set(os_list)&set(os_not_suuport)
                    if list(list_compare)==[]:
                        time.sleep(1)
                        self.driver.find_element(By.XPATH, "//input[@name='caseBO.actualWorkloadingArr']").send_keys(workloading_text)
                        time.sleep(1)
                        step_excute = self.driver.find_elements(By.XPATH,"//table[@id='excuteCaseTableRight']/tbody/tr[position()>1]/td/select")
                        for everystep in step_excute:
                            step = Select(everystep)
                            step.select_by_value('2')
                        self.driver.find_element(By.XPATH, "//input[@name='Save&Continue']").click()
                        time.sleep(2)
                        self.driver.close()
                        time.sleep(2)
                        self.driver.switch_to.window(main_window)
                        self.case_sheet.cell(column=long + 1, row=y, value="执行完成")
                    else:
                        time.sleep(1)
                        self.driver.find_element(By.XPATH, "//input[@name='caseBO.actualWorkloadingArr']").send_keys(workloading_text)
                        time.sleep(1)
                        os_number=len(os_list)
                        number_list=[x for x in range(os_number)]
                        for nosupportos in list(list_compare):
                            index=os_list.index(nosupportos)
                            number_list.remove(index)
                            td_position=(index+1)*2-1
                            step_excute = self.driver.find_elements(By.XPATH,f"//table[@id='excuteCaseTableRight']/tbody/tr[position()>1]/td[{td_position}]/select")
                            for everystep in step_excute:
                                step = Select(everystep)
                                step.select_by_value('7')
                                element_comment = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'planTestCaseItemResultComment')))
                                element_comment.send_keys('not support')
                                self.driver.find_element(By.XPATH, "//input[@name='OK' and @value='Save']").click()
                                time.sleep(1)
                        for pass_os in number_list:
                            td_position_pass = (pass_os + 1) * 2 - 1
                            step_excute_pass = self.driver.find_elements(By.XPATH, f"//table[@id='excuteCaseTableRight']/tbody/tr[position()>1]/td[{td_position_pass}]/select")
                            for everystep_pass in step_excute_pass:
                                step_pass = Select(everystep_pass)
                                step_pass.select_by_value('2')
                                time.sleep(1)
                        self.driver.find_element(By.XPATH, "//input[@name='Save&Continue']").click()
                        time.sleep(2)
                        self.driver.close()
                        time.sleep(2)
                        self.driver.switch_to.window(main_window)
                        self.case_sheet.cell(column=long + 1, row=y, value="执行完成")
                else:
                    os_list = []
                    self.driver.find_element(By.LINK_TEXT, every_row[0]).click()
                    popup_windows = self.driver.window_handles[-1]
                    self.driver.switch_to.window(popup_windows)
                    WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH, "//table[@id='excuteCaseTableRight']/thead/tr/td/strong/font")))
                    os = self.driver.find_elements(By.XPATH,"//table[@id='excuteCaseTableRight']/thead/tr/td/strong/font")
                    workloading = self.driver.find_element(By.XPATH, "//input[@id='caseForm_caseBO_workloading']")
                    workloading_text = workloading.get_attribute('value')
                    for everyos in os:
                        text = everyos.text
                        os_list.append(text)
                    list_compare = set(os_list) & set(os_not_suuport)
                    if list(list_compare) == []:
                        time.sleep(1)
                        self.driver.find_element(By.XPATH,"//input[@name='caseBO.actualWorkloadingArr']").send_keys(workloading_text)
                        time.sleep(1)
                        step_excute = self.driver.find_elements(By.XPATH,"//table[@id='excuteCaseTableRight']/tbody/tr[position()>1]/td/select")
                        first_step = Select(step_excute[0])
                        first_step.select_by_value('3')
                        element_comment = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'planTestCaseItemResultComment')))
                        self.driver.find_element(By.ID,"testCaseRelatedDefectId1").send_keys(self.project_issue_id)
                        self.driver.find_element(By.XPATH, "//tbody/tr/td/input[@value='Add' and @onclick='addTestCaseRelatedDefect()']").click()
                        time.sleep(2)
                        self.driver.find_element(By.ID, "planTestCaseItemResultComment")
                        self.driver.find_element(By.ID,"planTestCaseItemResultComment").clear()
                        del every_row[0]
                        for issue in every_row:
                            self.driver.find_element(By.ID, "planTestCaseItemResultComment").send_keys("TDMS#"+str(self.project_issue_id)+":" + issue + '\n')
                            time.sleep(1)
                        self.driver.find_element(By.XPATH,"//input[@name='OK']").click()
                        del step_excute[0]
                        for everystep in step_excute:
                            step = Select(everystep)
                            step.select_by_value('2')
                        self.driver.find_element(By.XPATH, "//input[@name='Save&Continue']").click()
                        time.sleep(2)
                        self.driver.close()
                        time.sleep(2)
                        self.driver.switch_to.window(main_window)
                        self.case_sheet.cell(column=long + 1, row=y, value="执行完成")
                    else:
                        time.sleep(1)
                        self.driver.find_element(By.XPATH,"//input[@name='caseBO.actualWorkloadingArr']").send_keys(workloading_text)
                        time.sleep(1)
                        os_number = len(os_list)
                        number_list = [x for x in range(os_number)]
                        for nosupportos in list(list_compare):
                            index = os_list.index(nosupportos)
                            number_list.remove(index)
                            td_position = (index + 1) * 2 - 1
                            step_excute = self.driver.find_elements(By.XPATH,f"//table[@id='excuteCaseTableRight']/tbody/tr[position()>1]/td[{td_position}]/select")
                            for everystep in step_excute:
                                step = Select(everystep)
                                step.select_by_value('7')
                                element_comment = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'planTestCaseItemResultComment')))
                                element_comment.send_keys('not support')
                                self.driver.find_element(By.XPATH, "//input[@name='OK' and @value='Save']").click()
                                time.sleep(1)
                        for pass_os in number_list:
                            td_position_pass = (pass_os + 1) * 2 - 1
                            step_excute_pass = self.driver.find_elements(By.XPATH,f"//table[@id='excuteCaseTableRight']/tbody/tr[position()>1]/td[{td_position_pass}]/select")
                            first_step = Select(step_excute_pass[0])
                            first_step.select_by_value('3')
                            element_comment = WebDriverWait(self.driver, 10000).until(ec.presence_of_element_located((By.ID, 'planTestCaseItemResultComment')))
                            self.driver.find_element(By.ID, "testCaseRelatedDefectId1").send_keys(self.project_issue_id)
                            self.driver.find_element(By.XPATH,"//tbody/tr/td/input[@value='Add' and @onclick='addTestCaseRelatedDefect()']").click()
                            time.sleep(1)
                            self.driver.find_element(By.ID, "planTestCaseItemResultComment")
                            self.driver.find_element(By.ID, "planTestCaseItemResultComment").clear()
                            copy_every_row=every_row
                            del copy_every_row[0]
                            for issue in copy_every_row:
                                self.driver.find_element(By.ID, "planTestCaseItemResultComment").send_keys("TDMS#"+str(self.project_issue_id)+":" + issue + '\n')
                                time.sleep(1)
                            self.driver.find_element(By.XPATH, "//input[@name='OK']").click()
                            del step_excute_pass[0]
                            for everystep_pass in step_excute_pass:
                                step_pass = Select(everystep_pass)
                                step_pass.select_by_value('2')
                                time.sleep(1)
                        self.driver.find_element(By.XPATH, "//input[@name='Save&Continue']").click()
                        time.sleep(3)
                        self.driver.close()
                        time.sleep(2)
                        self.driver.switch_to.window(main_window)
                        self.case_sheet.cell(column=long + 1, row=y, value="执行完成")
            except:
                self.driver.close()
                time.sleep(2)
                self.driver.switch_to.window(main_window)
                time.sleep(2)
                continue

    def run(self):
        self.enter_testcase()
        self.excel.save(r'C:\testplan.xlsx')

if __name__ == "__main__":
        case_go = auto_case()
        case_go.enter_plan()
        case_go.run()







